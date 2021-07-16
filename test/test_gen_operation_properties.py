from openpyxl import load_workbook
import yaml


property_prefix = "vesync.banner."
email_prefix = "push_banner_"
email_postfix = "@cloudtest.com"
country_code_to_timezone = {
    "US": "America/Los_Angeles",
    "CA": "America/Toronto",
    "DE": "Europe/Berlin",
    "AR": "America/Argentina/Buenos_Aires",
    "VN": "Asia/Ho_Chi_Minh"
}

wb = load_workbook("banner_accounts.xlsx")
ws = wb["accounts"]
property_list =[]
account_info_dict = {}

for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
    account_id = row[0].value
    phone_os = row[1].value
    app_version = row[2].value
    country_code = row[3].value

    if phone_os == "android":
        phone_os = "Android 6.0"
        app_version = f"VeSync {app_version}"
    elif phone_os == "ios":
        phone_os = "iOS 13.4"
        app_version = f"VeSync {app_version}"

    property_key_prefix = f"{property_prefix}account{account_id}."

    # property email
    property_email_key = f"{property_key_prefix}email"
    property_email_value = f"{email_prefix}{account_id}{email_postfix}"
    property_list.append(f"{property_email_key}={property_email_value}")

    # property phoneOS
    property_phone_os_key = f"{property_key_prefix}phoneOS"
    property_phone_os_value = phone_os
    property_list.append(f"{property_phone_os_key}={property_phone_os_value}")

    # property timezone
    property_timezone_key = f"{property_key_prefix}timeZone"
    property_timezone_value = country_code_to_timezone[country_code]
    property_list.append(f"{property_timezone_key}={property_timezone_value}")

    # property appVersion
    property_app_version_key = f"{property_key_prefix}appVersion"
    property_app_version_value = app_version
    property_list.append(f"{property_app_version_key}={property_app_version_value}")

    # property countryCode
    property_country_code_key = f"{property_key_prefix}countryCode"
    property_country_code_value = country_code
    property_list.append(f"{property_country_code_key}={property_country_code_value}")

    account_info_key = f"account{account_id}"
    account_info_value = {
        "email": f"${{__P({property_email_key})}}",
        "phoneOS": f"${{__P({property_phone_os_key})}}",
        "timeZone": f"${{__P({property_timezone_key})}}",
        "appVersion": f"${{__P({property_app_version_key})}}",
        "countryCode": f"${{__P({property_country_code_key})}}"
    }
    account_info_dict[account_info_key] = account_info_value


for p in property_list:
    print(p)

# print(yaml.dump(account_info_dict))
