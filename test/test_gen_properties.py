from  openpyxl import load_workbook


property_prefix = "mall.client."
email_prefix = "mall_client_user"
email_postfix = "@cloudtest.com"

wb = load_workbook("accounts.xlsx")
ws = wb["accounts"]
property_list =[]

for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
    account_id = row[0].value
    phone_os = row[1].value
    app_version = row[2].value
    guest_flag = row[3].value

    guest = None
    if guest_flag == 0:
        guest = False
    elif guest_flag == 1:
        guest = True

    if phone_os == "android":
        phone_os = "Android 6.0"
        app_version = f"VeSync {app_version}"
    elif phone_os == "ios":
        phone_os = "iOS 13.4"
        app_version = f"VeSync {app_version}"
    elif phone_os == "pc":
        phone_os = "PC"
        app_version = f"TestVersion"
    elif phone_os == "m":
        phone_os = "M"
        app_version = f"TestVersion"

    property_key_prefix = f"{property_prefix}account{account_id}."

    # property email
    property_email_key = f"{property_key_prefix}email"
    if guest:
        property_email_value = "''"
    else:
        property_email_value = f"{email_prefix}{account_id}{email_postfix}"
    property_list.append(f"{property_email_key}={property_email_value}")

    # property phoneOS
    property_phone_os_key = f"{property_key_prefix}phoneOS"
    property_phone_os_value = phone_os
    property_list.append(f"{property_phone_os_key}={property_phone_os_value}")

    # property timezone
    property_timezone_key = f"{property_key_prefix}timeZone"
    property_timezone_value = "America/Los_Angeles"
    property_list.append(f"{property_timezone_key}={property_timezone_value}")

    # property appVersion
    property_app_version_key = f"{property_key_prefix}appVersion"
    property_app_version_value = app_version
    property_list.append(f"{property_app_version_key}={property_app_version_value}")

for property in property_list:
    print(property)